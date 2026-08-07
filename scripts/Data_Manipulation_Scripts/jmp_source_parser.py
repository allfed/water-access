"""Parse JMP household country files for drinking-water source breakdown.

Adapted from the well-coverage repository (parse_jmp.py). Extracts urban/rural
percentages of population using piped, packaged, delivered, borehole/tubewell,
and other unpiped improved sources from the 'Water Data' sheet of each country
xlsx.

JMP template row indices (0-based, verified across GHA/BGD/IDN/KEN/MEX/NLD):
  6   All piped
  7   Non-piped (improved unpiped total)
  57  Tubewell, borehole
  61  Traditional wells
  73  All springs
  85  Rainwater
  88  Packaged water
  91  Surface water
  100 Cart with small tank/drum  (delivered water)
  101 Tanker truck provided      (delivered water)
  104 Other non-improved
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_MAX_YEAR_GAP = 5

DEFAULT_MANUAL_ALPHA3_MAPPING = {
    "Channel Islands": "CHI",
    "United States Virgin Islands": "VIR",
    "Wallis and Futuna Islands": "WLF",
    "Democratic Republic of the Congo": "COD",
    "China, Hong Kong SAR": "HKG",
    "China, Macao SAR": "MAC",
    "Republic of Korea": "KOR",
    "Curaçao": "CUW",
    "Sint Maarten (Dutch part)": "SXM",
    "Democratic People's Republic of Korea": "PRK",
}

# Top-level JMP rows used for scaling (avoid hierarchical double-count with piped/sub-rows)
BOREHOLE_ROWS = [57]
PACKAGED_ROWS = [88]
DELIVERED_ROWS = [100, 101]
OTHER_UNPIPED_ROWS = [61, 73, 85, 91, 104]

ROW_PIPED = 6
ROW_NON_PIPED = 7


def _resolve_raw_dir(raw_dir: Path | None = None) -> Path:
    """Return JMP raw xlsx directory, preferring explicit path then local then sibling repo."""
    if raw_dir is not None:
        return raw_dir
    repo_root = Path(__file__).resolve().parents[2]
    candidates = [
        repo_root / "data" / "original_data" / "jmp_country_files",
        repo_root.parent / "well-coverage" / "data" / "raw",
    ]
    for path in candidates:
        if path.is_dir() and any(path.glob("*.xlsx")):
            return path
    return candidates[0]


def survey_year(survey_id: str) -> int | None:
    match = re.search(r"_(\d{4})_", str(survey_id))
    return int(match.group(1)) if match else None


def _pct_value(rows: list[tuple], row_idx: int, col: int, offset: int) -> float | None:
    if row_idx >= len(rows):
        return None
    row = rows[row_idx]
    idx = col + offset
    if idx >= len(row):
        return None
    value = row[idx]
    if isinstance(value, (int, float)):
        return round(float(value), 4)
    return None


def _sum_pct(
    rows: list[tuple], row_indices: list[int], col: int, offset: int
) -> float | None:
    total = 0.0
    found = False
    for row_idx in row_indices:
        value = _pct_value(rows, row_idx, col, offset)
        if value is not None:
            total += value
            found = True
    return round(total, 4) if found else None


def _extract_block_pcts(rows: list[tuple], col: int) -> dict[str, float | None]:
    out: dict[str, float | None] = {}
    for suffix, offset in (("urban", 3), ("rural", 4), ("total", 5)):
        out[f"piped_{suffix}_pct"] = _pct_value(rows, ROW_PIPED, col, offset)
        out[f"non_piped_{suffix}_pct"] = _pct_value(rows, ROW_NON_PIPED, col, offset)
        out[f"borehole_{suffix}_pct"] = _sum_pct(rows, BOREHOLE_ROWS, col, offset)
        out[f"packaged_{suffix}_pct"] = _sum_pct(rows, PACKAGED_ROWS, col, offset)
        out[f"delivered_{suffix}_pct"] = _sum_pct(rows, DELIVERED_ROWS, col, offset)
        out[f"other_unpiped_{suffix}_pct"] = _sum_pct(
            rows, OTHER_UNPIPED_ROWS, col, offset
        )
    return out


def _survey_has_subcategory_data(record: dict[str, Any]) -> bool:
    keys = (
        "packaged_urban_pct",
        "packaged_rural_pct",
        "delivered_urban_pct",
        "delivered_rural_pct",
        "borehole_urban_pct",
        "borehole_rural_pct",
        "borehole_total_pct",
        "other_unpiped_urban_pct",
        "other_unpiped_rural_pct",
    )
    return any(record.get(key) is not None for key in keys)


def parse_country(path: Path) -> tuple[str, list[dict[str, Any]], str | None]:
    """Parse one JMP country workbook. Returns (iso3, survey records, error)."""
    iso3 = path.stem
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return iso3, [], f"load error: {exc}"

    if "Water Data" not in workbook.sheetnames:
        workbook.close()
        return iso3, [], "no Water Data sheet"

    rows = list(workbook["Water Data"].iter_rows(values_only=True))
    workbook.close()

    if len(rows) < 102:
        return iso3, [], "sheet too short"

    ids_row = rows[1]
    country = rows[0][3] if len(rows[0]) > 3 else iso3
    blocks = [
        (col, ids_row[col])
        for col in range(len(ids_row))
        if ids_row[col] and re.match(r"^[A-Z]{3}_\d{4}_", str(ids_row[col]))
    ]

    records: list[dict[str, Any]] = []
    for col, survey_id in blocks:
        year = survey_year(str(survey_id))
        pcts = _extract_block_pcts(rows, col)
        has_data = any(
            pcts.get(f"{key}_total_pct") is not None
            for key in ("packaged", "delivered", "borehole", "piped", "other_unpiped")
        )
        if not has_data:
            continue
        record = {
            "iso3": iso3,
            "country": country,
            "survey": survey_id,
            "year": year,
            "source_type": rows[2][col] if col < len(rows[2]) else None,
            **pcts,
        }
        records.append(record)

    return iso3, records, None


def _unpiped_budget(
    piped_pct: float | None, non_piped_pct: float | None
) -> float | None:
    if piped_pct is not None:
        return max(0.0, 100.0 - piped_pct)
    if non_piped_pct is not None:
        return max(0.0, non_piped_pct)
    return None


def normalize_unpiped_subcategories(
    *,
    piped_pct: float | None,
    non_piped_pct: float | None,
    packaged_pct: float | None,
    delivered_pct: float | None,
    borehole_pct: float | None,
    other_unpiped_pct: float | None,
) -> tuple[float, float, float, float]:
    """Scale tracked + other unpiped shares to fit the unpiped budget."""
    budget = _unpiped_budget(piped_pct, non_piped_pct)
    packaged = packaged_pct or 0.0
    delivered = delivered_pct or 0.0
    borehole = borehole_pct or 0.0
    other = other_unpiped_pct or 0.0
    total = packaged + delivered + borehole + other

    if budget is None:
        return packaged, delivered, borehole, other
    if total <= budget or total == 0:
        return packaged, delivered, borehole, other

    scale = budget / total
    return (
        round(packaged * scale, 4),
        round(delivered * scale, 4),
        round(borehole * scale, 4),
        round(other * scale, 4),
    )


def select_survey_for_country(
    records: list[dict[str, Any]],
    piped_reference_year: int | None,
    max_year_gap: int = DEFAULT_MAX_YEAR_GAP,
) -> dict[str, Any] | None:
    """Pick the most recent survey within max_year_gap of the piped reference year."""
    if not records:
        return None

    candidates = sorted(records, key=lambda r: r.get("year") or 0, reverse=True)
    if piped_reference_year is None:
        for rec in candidates:
            if _survey_has_subcategory_data(rec):
                return rec
        return None

    for rec in candidates:
        year = rec.get("year")
        if year is None:
            continue
        if abs(year - piped_reference_year) > max_year_gap:
            continue
        if _survey_has_subcategory_data(rec):
            return rec
    return None


def resolve_borehole_urban_rural(record: dict[str, Any]) -> tuple[float | None, float | None]:
    """Return urban/rural borehole percentages, falling back to total when needed."""
    urban = record.get("borehole_urban_pct")
    rural = record.get("borehole_rural_pct")
    if urban is None and rural is None:
        total = record.get("borehole_total_pct")
        if total is not None:
            return total, total
    return urban, rural


def apply_piped_reference_and_normalize(
    survey: dict[str, Any],
    urban_piped: float | None,
    rural_piped: float | None,
) -> dict[str, Any]:
    """Normalize unpiped subcategories against piped reference shares from merged data."""
    borehole_urban, borehole_rural = resolve_borehole_urban_rural(survey)

    pkg_u, del_u, bh_u, other_u = normalize_unpiped_subcategories(
        piped_pct=urban_piped,
        non_piped_pct=survey.get("non_piped_urban_pct"),
        packaged_pct=survey.get("packaged_urban_pct"),
        delivered_pct=survey.get("delivered_urban_pct"),
        borehole_pct=borehole_urban,
        other_unpiped_pct=survey.get("other_unpiped_urban_pct"),
    )
    pkg_r, del_r, bh_r, other_r = normalize_unpiped_subcategories(
        piped_pct=rural_piped,
        non_piped_pct=survey.get("non_piped_rural_pct"),
        packaged_pct=survey.get("packaged_rural_pct"),
        delivered_pct=survey.get("delivered_rural_pct"),
        borehole_pct=borehole_rural,
        other_unpiped_pct=survey.get("other_unpiped_rural_pct"),
    )

    return {
        **survey,
        "packaged_urban_pct": pkg_u if survey.get("packaged_urban_pct") is not None else 0.0,
        "packaged_rural_pct": pkg_r if survey.get("packaged_rural_pct") is not None else 0.0,
        "delivered_urban_pct": del_u if survey.get("delivered_urban_pct") is not None else 0.0,
        "delivered_rural_pct": del_r if survey.get("delivered_rural_pct") is not None else 0.0,
        "borehole_urban_pct": bh_u if borehole_urban is not None else 0.0,
        "borehole_rural_pct": bh_r if borehole_rural is not None else 0.0,
        "other_unpiped_urban_pct": other_u if survey.get("other_unpiped_urban_pct") is not None else 0.0,
        "other_unpiped_rural_pct": other_r if survey.get("other_unpiped_rural_pct") is not None else 0.0,
    }


def build_country_breakdown(
    all_records: list[dict[str, Any]],
    piped_reference: dict[str, dict[str, Any]] | None = None,
    max_year_gap: int = DEFAULT_MAX_YEAR_GAP,
) -> list[dict[str, Any]]:
    """Build per-country breakdown using year-filtered same-survey selection."""
    piped_reference = piped_reference or {}
    by_iso: dict[str, list[dict[str, Any]]] = {}
    for record in all_records:
        by_iso.setdefault(record["iso3"], []).append(record)

    breakdown_rows: list[dict[str, Any]] = []
    for iso3, records in sorted(by_iso.items()):
        ref = piped_reference.get(iso3, {})
        survey = select_survey_for_country(
            records,
            piped_reference_year=ref.get("piped_year"),
            max_year_gap=max_year_gap,
        )
        if survey is None:
            breakdown_rows.append({"iso3": iso3, "country": records[0].get("country")})
            continue

        normalized = apply_piped_reference_and_normalize(
            survey,
            urban_piped=ref.get("URBANPiped"),
            rural_piped=ref.get("RURALPiped"),
        )
        entry = {
            "iso3": iso3,
            "country": normalized.get("country"),
            "survey": normalized.get("survey"),
            "survey_year": normalized.get("year"),
            "piped_reference_year": ref.get("piped_year"),
        }
        entry.update(breakdown_to_model_columns(normalized))
        breakdown_rows.append(entry)

    return breakdown_rows


def breakdown_to_model_columns(row: dict[str, Any]) -> dict[str, float | None]:
    """Map parser output to merged_data column names."""
    borehole_urban, borehole_rural = resolve_borehole_urban_rural(row)
    return {
        "URBANPackaged": row.get("packaged_urban_pct"),
        "RURALPackaged": row.get("packaged_rural_pct"),
        "URBANDelivered": row.get("delivered_urban_pct"),
        "RURALDelivered": row.get("delivered_rural_pct"),
        "URBANBorehole": borehole_urban,
        "RURALBorehole": borehole_rural,
        "URBANOtherUnpiped": row.get("other_unpiped_urban_pct"),
        "RURALOtherUnpiped": row.get("other_unpiped_rural_pct"),
    }


def load_piped_reference_from_water_csv(
    water_csv_path: Path,
    manual_alpha3_mapping: dict[str, str] | None = None,
) -> dict[str, dict[str, Any]]:
    """Load latest piped year and urban/rural shares per alpha3 from the JMP/WHO CSV."""
    import pandas as pd
    import pycountry

    manual_alpha3_mapping = manual_alpha3_mapping or {}

    df = pd.read_csv(water_csv_path, low_memory=False)
    df.replace({"<1": 0, ">99": 100, "-": pd.NA}, inplace=True)
    for col in ("URBANPiped", "RURALPiped", "TOTALPiped"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    def _to_alpha3(country_name: str) -> str | None:
        if country_name in manual_alpha3_mapping:
            return manual_alpha3_mapping[country_name]
        cleaned = str(country_name).split("(")[0].strip()
        try:
            return pycountry.countries.search_fuzzy(cleaned)[0].alpha_3
        except LookupError:
            return None

    df["alpha3"] = df["Country"].map(_to_alpha3)
    df = df.sort_values(["Country", "Year"], ascending=[True, False])
    df = df.drop_duplicates("Country", keep="first")

    reference: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        alpha3 = row.get("alpha3")
        if not isinstance(alpha3, str) or not alpha3:
            continue
        reference[alpha3] = {
            "piped_year": int(row["Year"]) if pd.notna(row["Year"]) else None,
            "URBANPiped": float(row["URBANPiped"])
            if pd.notna(row.get("URBANPiped"))
            else None,
            "RURALPiped": float(row["RURALPiped"])
            if pd.notna(row.get("RURALPiped"))
            else None,
        }
    return reference


def parse_all_countries(
    raw_dir: Path | None = None,
    piped_reference: dict[str, dict[str, Any]] | None = None,
    max_year_gap: int = DEFAULT_MAX_YEAR_GAP,
) -> tuple[list[dict], list[tuple]]:
    """Parse every xlsx in raw_dir. Returns (country breakdown rows, errors)."""
    directory = _resolve_raw_dir(raw_dir)
    files = sorted(directory.glob("*.xlsx"))
    all_records: list[dict[str, Any]] = []
    errors: list[tuple[str, str]] = []

    for path in files:
        iso3, records, err = parse_country(path)
        if err:
            errors.append((iso3, err))
        else:
            all_records.extend(records)

    return (
        build_country_breakdown(
            all_records,
            piped_reference=piped_reference,
            max_year_gap=max_year_gap,
        ),
        errors,
    )
